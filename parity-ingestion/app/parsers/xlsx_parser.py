"""
Excel parsing — ported from backend/v1/parsing/xlsx_parser.py.
Produces backend-shaped row dicts, then ExtractionResult for the normaliser.
"""
from __future__ import annotations

import io
import math
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.models import ExtractionMethod, ExtractionResult, RawTransaction, WarningItem
from app.parsers.excel_common import canonical_hash, compute_txn_id, sort_rows
from app.parsing_errors import InvalidSchemaError
from app.xlsx_common import (
    REQUIRED_HEADERS,
    normalize_descriptor,
    normalize_header,
    parse_amount_with_detection,
    parse_date,
)

_EQUITY_EXCEL_COL_MAP = {
    "transaction date": "date",
    "transactio n date": "date",
    "transacti on date": "date",
    "transac tion date": "date",
    "transactiondate": "date",
    # December 2025+ layout: ignore secondary / merged-artifact columns
    "value date": None,
    "cheque number": "cheque_number",
    "cheque no": "cheque_number",
    "chq number": "cheque_number",
    "chq no": "cheque_number",
    "remarks 1": None,
    "remarks 2": None,
    "remarks": None,
    "narrative": "description",
    "details": "description",
    "transaction details": "description",
    "description": "description",
    "particulars": "description",
    "debit": "debit",
    "credit": "credit",
    "running balance": "balance",
    "balance": "balance",
    "transaction reference": "reference",
    "customer reference": "reference",
    "ref": "reference",
}


def _normalize_equity_col_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _is_blank_equity_header_cell(col: Any) -> bool:
    if col is None:
        return True
    if isinstance(col, float) and math.isnan(col):
        return True
    return False


EQUITY_HEADER_SCAN_MAX = 35


def _try_scan_equity_header_row(
    ws, max_scan: int = EQUITY_HEADER_SCAN_MAX
) -> Tuple[Optional[int], Optional[List[Any]]]:
    for row_num in range(1, max_scan + 1):
        row = next(ws.iter_rows(min_row=row_num, max_row=row_num), None)
        if row is None:
            break
        values = [cell.value for cell in row]
        if _is_equity_excel(values):
            return row_num, values
    return None, None


def scan_equity_excel_header(
    ws, *, max_scan: int = EQUITY_HEADER_SCAN_MAX
) -> Tuple[int, List[Any]]:
    """Return (zero-based row index, header values) for the first Equity-style header row."""
    row_num, values = _try_scan_equity_header_row(ws, max_scan)
    if row_num is None:
        raise InvalidSchemaError(
            f"Equity Excel header row not found in first {max_scan} rows"
        )
    return row_num - 1, values


def _is_equity_excel(header_row: List[Any]) -> bool:
    cols_lower = {
        _normalize_equity_col_name(c)
        for c in header_row
        if not _is_blank_equity_header_cell(c)
    }
    equity_indicators = {
        "narrative",
        "details",
        "transaction details",
        "running balance",
        "transactio n date",
        "transacti on date",
        "transac tion date",
        "transaction date",
    }
    if cols_lower & equity_indicators:
        return True
    # Nov/Dec 2025 exports: sometimes "Details" replaces "Narrative" — detect by bank columns.
    has_dc = "debit" in cols_lower and "credit" in cols_lower
    has_bal = "running balance" in cols_lower or "balance" in cols_lower
    has_desc = bool(
        cols_lower
        & {"narrative", "details", "transaction details", "description", "particulars"}
    )
    has_date = bool(
        cols_lower
        & {
            "transaction date",
            "transactio n date",
            "transacti on date",
            "transac tion date",
            "transactiondate",
        }
    )
    return has_dc and has_bal and has_desc and has_date


def _clean_equity_date_cell(value: Any) -> Any:
    """Strip embedded newlines from Equity date cells (e.g. '02-12-\\n2024')."""
    if value is None:
        return None
    if isinstance(value, str):
        return re.sub(r"[\r\n]+", "", value).strip()
    return value


_EQUITY_FOOTER_KEYWORDS = frozenset(
    {
        "total credits",
        "total debits",
        "closing balance",
        "opening balance",
        "total",
        "balance b/f",
        "balance c/f",
    }
)


def _equity_date_cell_looks_like_transaction(date_val: Any) -> bool:
    """
    Footer/summary rows leave the transaction date column blank. Real rows use
    datetime/date (typical in exports), Excel serials, or parseable date strings.
    """
    if date_val is None:
        return False
    if isinstance(date_val, str):
        s = date_val.strip()
        # Dec 2025 injects repeated column-header rows every ~25 data rows.
        # Gate them out before parse_date() is called. All three known
        # Equity date-header variants are listed here.
        if s.lower() in ('transaction date', 'transactio n date', 'transacti on date'):
            return False
        return bool(re.match(r'^\d{2}-\d{2}', s))
    if isinstance(date_val, (datetime, date)):
        return True
    if isinstance(date_val, bool):
        return False
    if isinstance(date_val, (int, float)):
        return True
    return False


def _equity_row_matches_footer_keywords(desc_val: Any, debit_val: Any, credit_val: Any) -> bool:
    """Bank footer lines often repeat labels in narrative or amount columns."""
    for v in (desc_val, debit_val, credit_val):
        if not isinstance(v, str):
            continue
        s = v.strip().lower()
        if not s:
            continue
        if s in _EQUITY_FOOTER_KEYWORDS:
            return True
        if s.startswith("total ") and ("credit" in s or "debit" in s):
            return True
    return False


def _equity_amount_cells_are_column_titles(debit_val: Any, credit_val: Any) -> bool:
    """
    Some Nov/Dec 2025 sheets repeat the column titles on the row immediately under
    the real header (cells contain the words 'Debit' / 'Credit', not amounts).
    Skip that row — it is not a transaction.
    """
    d = str(debit_val or "").strip().lower()
    c = str(credit_val or "").strip().lower()
    if d == "debit" and c == "credit":
        return True
    if d == "debit" and c in ("", "credit"):
        return True
    if c == "credit" and d in ("", "debit"):
        return True
    return False


def _normalise_equity_excel_columns(header_row: List[Any]) -> Dict[str, int]:
    """Build logical column index map; skips merged-cell None/NaN headers and ignored columns."""
    mapping: Dict[str, int] = {}
    for idx, col in enumerate(header_row):
        if _is_blank_equity_header_cell(col):
            continue
        key = _normalize_equity_col_name(col)
        if not key:
            continue
        if key not in _EQUITY_EXCEL_COL_MAP:
            continue
        mapped = _EQUITY_EXCEL_COL_MAP[key]
        if mapped is None:
            continue
        if mapped not in mapping:
            mapping[mapped] = idx
    required = {"date", "description"}
    missing = required - set(mapping.keys())
    if missing:
        raise InvalidSchemaError(f"Missing required Equity columns: {', '.join(sorted(missing))}")
    return mapping


def _header_map(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = normalize_header(cell)
        if not name:
            continue
        if name in mapping:
            raise InvalidSchemaError("Duplicate header detected")
        mapping[name] = idx
    missing = REQUIRED_HEADERS - set(mapping.keys())
    if missing:
        raise InvalidSchemaError(f"Missing required columns: {', '.join(sorted(missing))}")
    return mapping


def _detect_additional_header(second_row: List[Any]) -> bool:
    lowered = {normalize_header(c) for c in second_row if c not in (None, "")}
    return bool(REQUIRED_HEADERS & lowered)


def parse_xlsx(
    file_bytes: bytes, document_id: str, deal_currency: str
) -> Tuple[List[Dict[str, Any]], str, str, bool]:
    """
    Returns (rows_sorted, raw_hash, currency_detection, is_equity).
    """
    read_only = True
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=read_only)
    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        raise InvalidSchemaError("No visible worksheet found")
    if len(visible_sheets) > 1:
        raise InvalidSchemaError("Multiple visible worksheets are not allowed")

    ws = visible_sheets[0]

    scanned_num, scanned_header = _try_scan_equity_header_row(ws)
    if scanned_num is not None:
        header_row_num = scanned_num
        header_row = scanned_header
        is_equity = True
    else:
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_row_num = 1
        is_equity = _is_equity_excel(header_row)

    if not is_equity and not read_only:
        for merged in ws.merged_cells.ranges:
            if merged.min_row == 1:
                raise InvalidSchemaError("Merged cells in header row are not allowed")
    header_mapping = _normalise_equity_excel_columns(header_row) if is_equity else _header_map(header_row)

    second_row_cells = [
        cell.value
        for cell in next(ws.iter_rows(min_row=header_row_num + 1, max_row=header_row_num + 1), [])
    ]
    if second_row_cells and _detect_additional_header(second_row_cells):
        raise InvalidSchemaError("Multiple header rows detected")

    rows: List[Dict[str, Any]] = []
    currency_detection = "unknown"

    # header_row_num is 1-based (from _try_scan_equity_header_row); first data row is header_row_num + 1.
    for row in ws.iter_rows(min_row=header_row_num + 1):
        values = [cell.value for cell in row]
        if all(v in (None, "") for v in values):
            continue

        def get(col: str) -> Any:
            idx = header_mapping.get(col)
            return values[idx] if idx is not None and idx < len(values) else None

        date_val = get("date")
        txn_date_iso: Optional[str] = None
        if is_equity:
            date_val = _clean_equity_date_cell(date_val)
        desc_val = get("description")
        direction_val = get("direction") if "direction" in header_mapping else None

        debit_val = None
        credit_val = None
        if is_equity:
            debit_val = get("debit")
            credit_val = get("credit")
            if _equity_amount_cells_are_column_titles(debit_val, credit_val):
                continue
            if _equity_row_matches_footer_keywords(desc_val, debit_val, credit_val):
                continue
            reference_val = get("reference")
            if reference_val not in (None, ""):
                desc_val = f"{desc_val} {reference_val}".strip()
            cheque_num_val = get("cheque_number")
            if cheque_num_val not in (None, ""):
                desc_val = f"{desc_val} {cheque_num_val}".strip()
        else:
            amount_val = get("amount")

        if desc_val in (None, ""):
            if is_equity:
                continue
            raise InvalidSchemaError("Description is required per row")

        # Date guard — layout-agnostic.
        # Equity Excel stores real transaction dates as datetime objects (openpyxl).
        # Footer/summary rows (Total Credits, Total Debits, Closing Balance) leave
        # the date column blank (None) or non-datetime. Checking isinstance on the
        # raw cell value catches every footer variant regardless of wording, without
        # relying on keyword matching.
        # NOTE: only skip rows that fail this check; do NOT silence amount errors on
        # rows that DO have a valid datetime date.
        if is_equity:
            if not _equity_date_cell_looks_like_transaction(date_val):
                continue
            txn_date_iso = parse_date(date_val)

        try:
            if is_equity:
                debit_cents = 0
                credit_cents = 0
                if debit_val not in (None, ""):
                    debit_cents = abs(parse_amount_with_detection(debit_val, deal_currency)[0])
                if credit_val not in (None, ""):
                    credit_cents = abs(parse_amount_with_detection(credit_val, deal_currency)[0])
                signed_cents = credit_cents - debit_cents
                if signed_cents == 0:
                    raise InvalidSchemaError("Equity row missing both debit and credit")
            else:
                signed_cents, detection = parse_amount_with_detection(amount_val, deal_currency)
                if detection == "ambiguous":
                    currency_detection = "ambiguous"
        except InvalidSchemaError:
            raise
        except Exception as exc:
            raise InvalidSchemaError(str(exc)) from exc

        if direction_val:
            d = str(direction_val).strip().lower()
            if d in {"out", "debit", "withdrawal", "outflow"} and signed_cents > 0:
                signed_cents *= -1
            elif d in {"in", "credit", "inflow", "deposit"} and signed_cents < 0:
                signed_cents *= -1
            elif d not in {"out", "debit", "withdrawal", "outflow", "in", "credit", "inflow", "deposit"}:
                raise InvalidSchemaError(f"Invalid direction value: {direction_val}")

        txn_date = txn_date_iso if is_equity else parse_date(date_val)

        row_obj: Dict[str, Any] = {
            "txn_date": txn_date,
            "signed_amount_cents": signed_cents,
            "abs_amount_cents": abs(signed_cents),
            "raw_descriptor": str(desc_val),
            "parsed_descriptor": str(desc_val).strip(),
            "normalized_descriptor": normalize_descriptor(desc_val),
            "account_id": "default",
        }
        row_obj["txn_id"] = compute_txn_id(row_obj, document_id)
        rows.append(row_obj)

    wb.close()

    if not rows:
        raise InvalidSchemaError("Sheet has no data rows")

    if is_equity:
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                r["txn_date"],
                r["account_id"],
                r["signed_amount_cents"],
                r["normalized_descriptor"],
                r["txn_id"],
            ),
        )
    else:
        rows_sorted = sort_rows(rows)
    raw_hash = canonical_hash(rows_sorted)
    return rows_sorted, raw_hash, currency_detection, is_equity


def _signed_cents_to_debit_credit_raw(signed_cents: int) -> tuple[str, str]:
    ac = abs(signed_cents)
    s = f"{ac // 100}.{ac % 100:02d}"
    if signed_cents > 0:
        return "", s
    if signed_cents < 0:
        return s, ""
    return "", ""


def extraction_result_from_xlsx_bytes(
    file_bytes: bytes,
    document_id: str,
    deal_currency: str,
    source_file: str,
) -> Tuple[ExtractionResult, str]:
    """
    Parse XLSX bytes to ExtractionResult + raw_transaction_hash (same pipeline as PDF).
    """
    rows_sorted, raw_hash, currency_detection, is_equity = parse_xlsx(
        file_bytes, document_id, deal_currency
    )
    ext_type: ExtractionMethod = "equity_xlsx" if is_equity else "xlsx_generic"
    cur = "KES"
    if currency_detection == "ambiguous":
        cur = "KES"

    raw_transactions: List[RawTransaction] = []
    for i, row in enumerate(rows_sorted):
        sc = int(row["signed_amount_cents"])
        dr, cr = _signed_cents_to_debit_credit_raw(sc)
        raw_transactions.append(
            RawTransaction(
                row_index=i,
                date_raw=str(row["txn_date"]),
                description=str(row["raw_descriptor"]),
                debit_raw=dr,
                credit_raw=cr,
                balance_raw="",
                source_file=source_file,
                extraction_confidence=1.0,
                source_extraction_method="equity_xlsx" if is_equity else "xlsx_generic",
            )
        )

    result = ExtractionResult(
        source_file=source_file,
        extractor_type=ext_type,
        row_count=len(raw_transactions),
        extraction_status="success",
        warnings=[],
        raw_transactions=raw_transactions,
        currency=cur,
    )
    return result, raw_hash
