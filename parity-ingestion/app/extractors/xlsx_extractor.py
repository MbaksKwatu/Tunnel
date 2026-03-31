"""
XLSX bank / generic spreadsheet extraction for parity-ingestion.

Ported from backend/v1/parsing/xlsx_parser.py — returns ExtractionResult with
RawTransaction rows (same shape as PDF extractors). Uses read_only workbook
to limit memory on large files.
"""
from __future__ import annotations

import re
from typing import Any, Dict, List

from openpyxl import load_workbook

from app.models import ExtractionResult, RawTransaction, WarningItem
from app.parsing_errors import InvalidSchemaError
from app.xlsx_common import (
    OPTIONAL_HEADERS,
    REQUIRED_HEADERS,
    normalize_descriptor,
    normalize_header,
    parse_amount_with_detection,
    parse_date,
)

_EQUITY_EXCEL_COL_MAP = {
    "transaction date": "date",
    "transactio n date": "date",
    "transactiondate": "date",
    "value date": "value_date",
    "narrative": "description",
    "particulars": "description",
    "debit": "debit",
    "credit": "credit",
    "running balance": "balance",
    "balance": "balance",
    "transaction reference": "reference",
    "customer reference": "reference",
    "ref": "reference",
}


def _normalize_equity_col_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _is_equity_excel(header_row: List[Any]) -> bool:
    cols_lower = {_normalize_equity_col_name(c) for c in header_row if c is not None}
    equity_indicators = {"narrative", "running balance", "transactio n date", "transaction date"}
    return bool(cols_lower & equity_indicators)


def _normalise_equity_excel_columns(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, col in enumerate(header_row):
        key = _normalize_equity_col_name(col)
        if not key:
            continue
        mapped = _EQUITY_EXCEL_COL_MAP.get(key)
        if not mapped:
            continue
        if mapped not in mapping:
            mapping[mapped] = idx
    required = {"date", "description"}
    missing = required - set(mapping.keys())
    if missing:
        raise InvalidSchemaError(f"Missing required Equity columns: {', '.join(sorted(missing))}")
    return mapping


def _header_map(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = normalize_header(cell)
        if not name:
            continue
        if name in mapping:
            raise InvalidSchemaError("Duplicate header detected")
        mapping[name] = idx
    missing = REQUIRED_HEADERS - set(mapping.keys())
    if missing:
        raise InvalidSchemaError(f"Missing required columns: {', '.join(sorted(missing))}")
    return mapping


def _detect_additional_header(second_row: List[Any]) -> bool:
    lowered = {normalize_header(c) for c in second_row if c not in (None, "")}
    return bool(REQUIRED_HEADERS & lowered)


def _cell_amount_str(val: Any) -> str:
    if val in (None, ""):
        return ""
    if isinstance(val, (int, float)):
        c = int(round(float(val) * 100))
        neg = c < 0
        c = abs(c)
        return f"{'-' if neg else ''}{c // 100}.{c % 100:02d}"
    return str(val).strip()


def extract_xlsx(file_path: str, *, deal_currency: str = "KES") -> ExtractionResult:
    """
    Extract transactions from a single-sheet XLSX.

    Equity Bank exports and generic date/amount/description spreadsheets are supported.
    """
    read_only = True
    wb = load_workbook(file_path, data_only=True, read_only=read_only)
    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        raise InvalidSchemaError("No visible worksheet found")
    if len(visible_sheets) > 1:
        raise InvalidSchemaError("Multiple visible worksheets are not allowed")

    ws = visible_sheets[0]

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    is_equity = _is_equity_excel(header_row)
    header_mapping = _normalise_equity_excel_columns(header_row) if is_equity else _header_map(header_row)

    second_row_cells = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2), [])]
    if second_row_cells and _detect_additional_header(second_row_cells):
        raise InvalidSchemaError("Multiple header rows detected")

    raw_transactions: List[RawTransaction] = []
    warnings: List[WarningItem] = []
    currency_detection = "unknown"
    row_idx = 0

    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if all(v in (None, "") for v in values):
            continue

        def get(col: str) -> Any:
            idx = header_mapping.get(col)
            return values[idx] if idx is not None and idx < len(values) else None

        date_val = get("date")
        desc_val = get("description")
        direction_val = get("direction") if "direction" in header_mapping else None

        debit_val = None
        credit_val = None
        balance_val = None
        if is_equity:
            debit_val = get("debit")
            credit_val = get("credit")
            balance_val = get("balance")
            reference_val = get("reference")
            if reference_val not in (None, ""):
                desc_val = f"{desc_val} {reference_val}".strip()
        else:
            amount_val = get("amount")

        if desc_val in (None, ""):
            if is_equity:
                continue
            raise InvalidSchemaError("Description is required per row")

        debit_raw = ""
        credit_raw = ""
        balance_raw = ""
        signed_cents = 0

        try:
            if is_equity:
                debit_cents = 0
                credit_cents = 0
                if debit_val not in (None, ""):
                    debit_cents = abs(parse_amount_with_detection(debit_val, deal_currency)[0])
                if credit_val not in (None, ""):
                    credit_cents = abs(parse_amount_with_detection(credit_val, deal_currency)[0])
                signed_cents = credit_cents - debit_cents
                if signed_cents == 0:
                    raise InvalidSchemaError("Equity row missing both debit and credit")
                debit_raw = _cell_amount_str(debit_val) if debit_val not in (None, "") else ""
                credit_raw = _cell_amount_str(credit_val) if credit_val not in (None, "") else ""
                balance_raw = _cell_amount_str(balance_val) if balance_val not in (None, "") else ""
            else:
                signed_cents, detection = parse_amount_with_detection(amount_val, deal_currency)
                if detection == "ambiguous":
                    currency_detection = "ambiguous"
                if direction_val:
                    d = str(direction_val).strip().lower()
                    if d in {"out", "debit", "withdrawal", "outflow"} and signed_cents > 0:
                        signed_cents *= -1
                    elif d in {"in", "credit", "inflow", "deposit"} and signed_cents < 0:
                        signed_cents *= -1
                    elif d not in {"out", "debit", "withdrawal", "outflow", "in", "credit", "inflow", "deposit"}:
                        raise InvalidSchemaError(f"Invalid direction value: {direction_val}")
                ac = abs(signed_cents)
                amt_display = f"{ac // 100}.{ac % 100:02d}"
                if signed_cents > 0:
                    credit_raw = amt_display
                elif signed_cents < 0:
                    debit_raw = amt_display
        except InvalidSchemaError:
            raise
        except Exception as exc:
            raise InvalidSchemaError(str(exc)) from exc

        parse_date(date_val)  # validate — same rules as backend xlsx_parser
        date_raw_str = str(date_val).strip() if date_val is not None else ""

        raw_transactions.append(
            RawTransaction(
                row_index=row_idx,
                date_raw=date_raw_str,
                description=str(desc_val).strip(),
                debit_raw=debit_raw,
                credit_raw=credit_raw,
                balance_raw=balance_raw,
                source_file=file_path,
                extraction_confidence=1.0,
                source_extraction_method="equity_xlsx" if is_equity else "xlsx_generic",
            )
        )
        row_idx += 1

    wb.close()

    if not raw_transactions:
        raise InvalidSchemaError("Sheet has no data rows")

    ext_type: Any = "equity_xlsx" if is_equity else "xlsx_generic"
    cur = "KES"
    if currency_detection == "ambiguous":
        cur = "KES"

    return ExtractionResult(
        source_file=file_path,
        extractor_type=ext_type,
        row_count=len(raw_transactions),
        extraction_status="success",
        warnings=warnings,
        raw_transactions=raw_transactions,
        currency=cur,
    )
