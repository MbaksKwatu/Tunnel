import io
import re
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from .common import (
    REQUIRED_HEADERS,
    OPTIONAL_HEADERS,
    canonical_hash,
    compute_txn_id,
    normalize_descriptor,
    normalize_header,
    parse_amount_with_detection,
    parse_date,
    sort_rows,
)
from .errors import InvalidSchemaError

_EQUITY_EXCEL_COL_MAP = {
    "transaction date": "date",
    "transactio n date": "date",
    "transactiondate": "date",
    "value date": "value_date",
    "narrative": "description",
    "particulars": "description",
    "debit": "debit",
    "credit": "credit",
    "running balance": "balance",
    "balance": "balance",
    "transaction reference": "reference",
    "customer reference": "reference",
    "ref": "reference",
}


def _normalize_equity_col_name(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _is_equity_excel(header_row: List[Any]) -> bool:
    cols_lower = {_normalize_equity_col_name(c) for c in header_row if c is not None}
    equity_indicators = {"narrative", "running balance", "transactio n date", "transaction date"}
    return bool(cols_lower & equity_indicators)


def _normalise_equity_excel_columns(header_row: List[Any]) -> Dict[str, int]:
    """
    Map Equity-exported headers to internal parser columns.
    Keeps the first index per normalized name and ignores empty/None headers.
    """
    mapping: Dict[str, int] = {}
    for idx, col in enumerate(header_row):
        key = _normalize_equity_col_name(col)
        if not key:
            continue
        mapped = _EQUITY_EXCEL_COL_MAP.get(key)
        if not mapped:
            continue
        if mapped not in mapping:
            mapping[mapped] = idx
    required = {"date", "description"}
    missing = required - set(mapping.keys())
    if missing:
        raise InvalidSchemaError(f"Missing required Equity columns: {', '.join(sorted(missing))}")
    return mapping


def _header_map(header_row: List[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = normalize_header(cell)
        if not name:
            continue
        if name in mapping:
            raise InvalidSchemaError("Duplicate header detected")
        mapping[name] = idx
    missing = REQUIRED_HEADERS - set(mapping.keys())
    if missing:
        raise InvalidSchemaError(f"Missing required columns: {', '.join(sorted(missing))}")
    return mapping


def _detect_additional_header(second_row: List[Any]) -> bool:
    # If second row still contains any required header tokens, treat as multiple header rows
    lowered = {normalize_header(c) for c in second_row if c not in (None, "")}
    return bool(REQUIRED_HEADERS & lowered)


def parse_xlsx(file_bytes: bytes, document_id: str, deal_currency: str) -> Tuple[List[Dict[str, Any]], str, str]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        raise InvalidSchemaError("No visible worksheet found")
    if len(visible_sheets) > 1:
        raise InvalidSchemaError("Multiple visible worksheets are not allowed")

    ws = visible_sheets[0]

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    is_equity = _is_equity_excel(header_row)
    # Keep strict merged-header rejection for generic spreadsheets.
    # Equity exports commonly contain merged header artifacts that produce None columns.
    if not is_equity:
        for merged in ws.merged_cells.ranges:
            if merged.min_row == 1:
                raise InvalidSchemaError("Merged cells in header row are not allowed")
    header_mapping = _normalise_equity_excel_columns(header_row) if is_equity else _header_map(header_row)

    # Optional: detect multiple header rows
    second_row_cells = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2), [])]
    if second_row_cells and _detect_additional_header(second_row_cells):
        raise InvalidSchemaError("Multiple header rows detected")

    rows: List[Dict[str, Any]] = []
    currency_detection = "unknown"

    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if all(v in (None, "") for v in values):
            continue  # skip empty rows

        def get(col: str) -> Any:
            idx = header_mapping.get(col)
            return values[idx] if idx is not None and idx < len(values) else None

        date_val = get("date")
        desc_val = get("description")
        direction_val = get("direction") if "direction" in header_mapping else None

        debit_val = None
        credit_val = None
        if is_equity:
            debit_val = get("debit")
            credit_val = get("credit")
            reference_val = get("reference")
            if reference_val not in (None, ""):
                desc_val = f"{desc_val} {reference_val}".strip()
        else:
            amount_val = get("amount")

        if desc_val in (None, ""):
            if is_equity:
                # Equity exports include non-transaction footer rows (e.g., end-of-statement markers).
                continue
            raise InvalidSchemaError("Description is required per row")

        # Amount parsing and currency detection
        try:
            if is_equity:
                debit_cents = 0
                credit_cents = 0
                if debit_val not in (None, ""):
                    debit_cents = abs(parse_amount_with_detection(debit_val, deal_currency)[0])
                if credit_val not in (None, ""):
                    credit_cents = abs(parse_amount_with_detection(credit_val, deal_currency)[0])
                signed_cents = credit_cents - debit_cents
                if signed_cents == 0:
                    raise InvalidSchemaError("Equity row missing both debit and credit")
            else:
                signed_cents, detection = parse_amount_with_detection(amount_val, deal_currency)
                if detection == "ambiguous":
                    currency_detection = "ambiguous"
        except InvalidSchemaError:
            raise
        except Exception as exc:
            raise InvalidSchemaError(str(exc)) from exc

        # Direction handling
        if direction_val:
            d = str(direction_val).strip().lower()
            if d in {"out", "debit", "withdrawal", "outflow"} and signed_cents > 0:
                signed_cents *= -1
            elif d in {"in", "credit", "inflow", "deposit"} and signed_cents < 0:
                signed_cents *= -1
            elif d not in {"out", "debit", "withdrawal", "outflow", "in", "credit", "inflow", "deposit"}:
                raise InvalidSchemaError(f"Invalid direction value: {direction_val}")

        txn_date = parse_date(date_val)

        row_obj: Dict[str, Any] = {
            "txn_date": txn_date,
            "signed_amount_cents": signed_cents,
            "abs_amount_cents": abs(signed_cents),
            "raw_descriptor": str(desc_val),
            "parsed_descriptor": str(desc_val).strip(),
            "normalized_descriptor": normalize_descriptor(desc_val),
            "account_id": "default",
        }
        row_obj["txn_id"] = compute_txn_id(row_obj, document_id)
        rows.append(row_obj)

    if not rows:
        raise InvalidSchemaError("Sheet has no data rows")

    if is_equity:
        # Preserve all Equity Excel rows (no structural dedupe) to match statement row counts.
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                r["txn_date"],
                r["account_id"],
                r["signed_amount_cents"],
                r["normalized_descriptor"],
                r["txn_id"],
            ),
        )
    else:
        rows_sorted = sort_rows(rows)
    raw_hash = canonical_hash(rows_sorted)
    return rows_sorted, raw_hash, currency_detection
