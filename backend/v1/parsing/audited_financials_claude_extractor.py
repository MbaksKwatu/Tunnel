"""
Claude-based audited financials / management accounts extractor.

Isolated extraction path: takes one file (PDF — text or scanned, image,
XLSX, or CSV) and asks Claude to read it natively (PDF/vision input for
PDFs and images; the raw text/cells for spreadsheets) and return the same
field schema as the existing pdfplumber-coordinate extractor in
parity-ingestion/app/extractors/audited_financials_extractor.py.

Architecture decision and rationale: see
Tunnel/docs/AUDITED_FINANCIALS_EXTRACTION_INVESTIGATION.md, section 8.

This module is intentionally standalone — it does not import from, call,
or modify the existing extractor, the parity-ingestion client, or the
inline fallback. It is not wired into any ingestion endpoint and writes
nothing to the database. No fallback model, no retry-with-different-model
logic, no confidence-score reproduction.
"""
from __future__ import annotations

import base64
import csv
import io
import json
import logging
import os
from pathlib import Path
from typing import Any, Dict

import anthropic

logger = logging.getLogger(__name__)

# Pinned for every file and every format (PDF, image, spreadsheet) this module
# handles. Do not branch on file type/extension to pick a different model —
# document-layout heterogeneity is already the main source of variation across
# the validation corpus; per-format model switching would confound it with a
# second, uncontrolled variable. See AUDITED_FINANCIALS_EXTRACTION_INVESTIGATION.md
# section 11.
MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 8192

_PDF_EXTENSIONS = {".pdf"}
_IMAGE_MIME = {
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".webp": "image/webp",
    ".gif": "image/gif",
}
_SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}

_SCHEMA_INSTRUCTIONS = """\
You are extracting structured financial data from an audited financial \
statement or management account document. Read the document directly — \
tables, columns, and notes — and return ONLY a single JSON object (no \
prose, no markdown fences) with exactly these fields. Use the MOST RECENT \
financial year's figures (not prior-year comparatives) for every amount. \
All monetary fields are integers in CENTS of the document's stated \
currency (e.g. KES 1,234,567.00 -> 123456700). Use null for any field the \
document does not state — never guess, never substitute zero for missing.

{
  "company_name": string or null,
  "financial_year": integer (e.g. 2025) or null,
  "financial_year_start": "YYYY-MM-DD" or null,
  "financial_year_end": "YYYY-MM-DD" or null,
  "currency": string or null (ISO-4217-style code if determinable, e.g. "KES", "USD"; null if not stated anywhere in the document),
  "auditor_name": string or null,

  "turnover_cents": integer or null (revenue / turnover / total income, current year),
  "other_income_cents": integer or null,
  "cost_of_sales_cents": integer or null (positive magnitude),
  "operating_costs_cents": integer or null (positive magnitude),
  "administrative_costs_cents": integer or null (positive magnitude),
  "staff_costs_cents": integer or null (positive magnitude),
  "finance_costs_cents": integer or null (positive magnitude),
  "depreciation_expense_cents": integer or null (positive magnitude),
  "other_expenses_cents": integer or null (positive magnitude),
  "total_expenses_cents": integer or null (the income statement's OWN stated total expenses / total operating expenses / total costs line, current year, positive magnitude — copy the figure the document actually prints on that total line; do NOT add up the sub-category fields above yourself, and return null if the document shows no such total line),
  "profit_before_tax_cents": integer or null,
  "tax_expense_cents": integer or null (positive magnitude),
  "profit_after_tax_cents": integer or null,

  "property_plant_equipment_cents": integer or null,
  "intangible_assets_cents": integer or null,
  "investments_cents": integer or null,
  "inventory_cents": integer or null,
  "trade_receivables_cents": integer or null,
  "other_receivables_cents": integer or null,
  "prepayments_cents": integer or null,
  "cash_and_equivalents_cents": integer or null (total cash & bank balances, balance sheet figure),
  "total_assets_cents": integer or null,
  "trade_payables_cents": integer or null,
  "tax_payable_cents": integer or null,
  "short_term_loans_cents": integer or null,
  "long_term_loans_cents": integer or null,
  "other_payables_cents": integer or null,
  "other_noncurrent_liabilities_cents": integer or null,
  "share_capital_cents": integer or null,
  "retained_earnings_cents": integer or null,
  "other_reserves_cents": integer or null,
  "equity_cents": integer or null (total equity),
  "total_equity_and_liabilities_cents": integer or null,

  "operating_cashflow_cents": integer or null (net cash from/used in operating activities),
  "investing_cashflow_cents": integer or null (net cash from/used in investing activities),
  "financing_cashflow_cents": integer or null (net cash from/used in financing activities),
  "cash_at_start_cents": integer or null,
  "cash_at_end_cents": integer or null,

  "cash_breakdown": object or null (per-bank-account cash balances at year-end from the cash/bank notes, e.g. {"KCB": 3031250000, "Equity": 15558720000}; cents; null if the document has no such breakdown, even if cash_and_equivalents_cents is present),

  "loan_breakdown": array or null (one entry per loan/borrowing facility disclosed in the loan note, e.g. Note 14 or equivalent; null if no such note exists). Each entry:
    {
      "name": string (lender/facility name as stated, e.g. "KCB Asset Finance"),
      "reference": string or null (facility/account/loan reference number as stated, if disclosed),
      "type": string or null (facility type as stated, e.g. "Term Loan", "Overdraft", "Asset Finance", "Mortgage"),
      "amount_cents": integer (outstanding balance for this facility, current year, cents)
    }
}

Field-extraction rules:
- Expense fields (cost_of_sales_cents, operating_costs_cents, administrative_costs_cents, \
staff_costs_cents, finance_costs_cents, depreciation_expense_cents, other_expenses_cents, \
tax_expense_cents) must be positive magnitudes even if the document shows them in \
parentheses or as negative.
- total_expenses_cents must be the single total-expenses figure the income statement itself \
prints (e.g. a "Total expenses", "Total operating expenses", or "Total costs" line). Transcribe \
that stated total verbatim; do not derive it by summing cost_of_sales_cents, operating_costs_cents, \
etc. If no such total line is printed anywhere, return null — never substitute a self-computed sum.
- cash_breakdown and loan_breakdown come from the notes to the accounts (the cash/bank note \
and the borrowings/loans note), not from the face of the balance sheet — only populate them \
if the document actually contains such a note with a breakdown.
- If the document is a CSV/XLSX export rather than a formatted statement, apply the same \
field mapping to whatever labeled rows/columns are present.
- Return null for any field not present anywhere in the document. Do not fabricate, \
interpolate, or carry forward a prior-year value into a current-year field.
"""


# total_liabilities_cents is intentionally not part of _SCHEMA_INSTRUCTIONS —
# the model was previously asked to compute this total itself, and in the
# 6-file validation it omitted long-term loans in 4 of 6 files despite
# correctly extracting them into long_term_loans_cents in the same response
# (see AUDITED_FINANCIALS_EXTRACTION_INVESTIGATION.md section 9.5, finding 1).
# It is now computed deterministically from the granular liability fields
# the model does extract reliably.
_LIABILITY_COMPONENT_FIELDS = (
    "trade_payables_cents",
    "tax_payable_cents",
    "short_term_loans_cents",
    "long_term_loans_cents",
    "other_payables_cents",
    "other_noncurrent_liabilities_cents",
)


def _compute_total_liabilities_cents(data: Dict[str, Any]) -> int | None:
    """Sum the granular liability fields. None if the model found none of them.

    Deliberately does not guess at liability components that have no
    corresponding schema field (e.g. a "due to related parties" balance, or
    deferred income shown under current liabilities) — a document with such
    a component will produce a total that undercounts real total liabilities
    rather than one that silently looks complete. Callers that need to know
    whether that gap exists should check the document directly.
    """
    components = [data.get(field) for field in _LIABILITY_COMPONENT_FIELDS]
    populated = [c for c in components if c is not None]
    if not populated:
        return None
    return sum(populated)


class ClaudeExtractionError(Exception):
    """Raised when Claude cannot read or extract from the document at all."""


def _media_type_and_kind(file_name: str) -> tuple[str, str]:
    """Return (kind, media_type) where kind is one of: pdf, image, spreadsheet."""
    ext = Path(file_name).suffix.lower()
    if ext in _PDF_EXTENSIONS:
        return "pdf", "application/pdf"
    if ext in _IMAGE_MIME:
        return "image", _IMAGE_MIME[ext]
    if ext in _SPREADSHEET_EXTENSIONS:
        return "spreadsheet", ""
    raise ClaudeExtractionError(f"Unsupported file extension: '{ext}'")


def _spreadsheet_to_text(file_bytes: bytes, file_name: str) -> str:
    """Render a CSV/XLSX file as plain text (one row per line) for a text content block."""
    ext = Path(file_name).suffix.lower()
    if ext == ".csv":
        text = file_bytes.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        rows = []
        for sheet in wb.worksheets:
            rows.append([f"--- sheet: {sheet.title} ---"])
            for row in sheet.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in row])

    return "\n".join(", ".join(cell for cell in row) for row in rows)


def _build_content_block(file_bytes: bytes, file_name: str) -> Dict[str, Any]:
    kind, media_type = _media_type_and_kind(file_name)

    if kind == "pdf":
        return {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.standard_b64encode(file_bytes).decode("utf-8"),
            },
        }
    if kind == "image":
        return {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_type,
                "data": base64.standard_b64encode(file_bytes).decode("utf-8"),
            },
        }
    # spreadsheet — no native document block; send as text
    text = _spreadsheet_to_text(file_bytes, file_name)
    return {"type": "text", "text": text}


def extract_audited_financials_claude(
    file_bytes: bytes,
    file_name: str,
) -> Dict[str, Any]:
    """
    Extract structured financial data from a single audited-financials or
    management-account file using Claude as the extraction engine.

    `file_bytes` / `file_name` are the raw upload — same shape as every
    other extractor in this codebase. Returns a dict matching the field
    schema documented in _SCHEMA_INSTRUCTIONS above (a superset of the
    existing pdfplumber-coordinate extractor's output shape).

    Raises ClaudeExtractionError if the API call fails, the file type is
    unsupported, or the response is not parseable JSON. Does not retry,
    does not fall back to another model, does not compute a confidence
    score.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise ClaudeExtractionError("ANTHROPIC_API_KEY is not set")

    try:
        content_block = _build_content_block(file_bytes, file_name)
    except ClaudeExtractionError:
        raise
    except Exception as exc:
        raise ClaudeExtractionError(f"Could not prepare '{file_name}' for extraction: {exc}") from exc

    client = anthropic.Anthropic(api_key=api_key)

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            messages=[
                {
                    "role": "user",
                    "content": [
                        content_block,
                        {"type": "text", "text": _SCHEMA_INSTRUCTIONS},
                    ],
                }
            ],
        )
    except anthropic.APIError as exc:
        raise ClaudeExtractionError(f"Claude API call failed for '{file_name}': {exc}") from exc

    if response.stop_reason == "max_tokens":
        raise ClaudeExtractionError(
            f"Claude response for '{file_name}' was truncated at max_tokens={MAX_TOKENS}"
        )

    text = "".join(block.text for block in response.content if block.type == "text").strip()
    # Claude occasionally wraps the JSON in a markdown fence and/or a leading
    # sentence ("Here is the extracted JSON:") despite being told not to.
    # Take the substring between the first '{' and the matching last '}'.
    start = text.find("{")
    end = text.rfind("}")
    if start != -1 and end != -1 and end > start:
        text = text[start : end + 1]

    try:
        data: Dict[str, Any] = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ClaudeExtractionError(
            f"Claude response for '{file_name}' was not valid JSON: {exc}\nRaw response: {text[:2000]}"
        ) from exc

    data["total_liabilities_cents"] = _compute_total_liabilities_cents(data)
    data["extraction_method"] = "claude_sonnet_4_6"
    data["_usage"] = {
        "input_tokens": response.usage.input_tokens,
        "output_tokens": response.usage.output_tokens,
    }

    logger.info(
        "[CLAUDE AUDITED] Extracted %s FY%s from '%s' — input_tokens=%d output_tokens=%d",
        data.get("company_name"),
        data.get("financial_year"),
        file_name,
        response.usage.input_tokens,
        response.usage.output_tokens,
    )
    return data


__all__ = ["extract_audited_financials_claude", "ClaudeExtractionError"]
