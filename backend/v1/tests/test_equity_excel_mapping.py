from __future__ import annotations

import datetime
import io
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from v1.parsing.xlsx_parser import (
    _is_equity_excel,
    _normalise_equity_excel_columns,
    parse_xlsx,
    scan_equity_excel_header,
)

# Real-file Jan/Feb tests were removed: the committed synthetic fixtures in
# tests_v1/fixtures/ (test_equity_jan_fixture.xlsx, test_equity_feb_fixture.xlsx)
# cover the same column-mapping and sign logic in test_parsers.py::TestDeterministicParsers.
# Add a real large-scale file test here only if you commit the XLSX to tests_v1/fixtures/.


def _december_2025_layout_workbook_bytes() -> bytes:
    """Synthetic December 2025 Equity layout: preamble, header row 7, 3030 data rows, 2 footer rows."""
    wb = Workbook()
    ws = wb.active
    for _ in range(5):
        ws.append(["preamble", "", "", "", ""])
    ws.append(["", "", "", "", "Total Count: 3030"])
    header = [
        "Transaction Date",
        "Value Date",
        None,
        "Narrative",
        "Extra",
        None,
        "Debit",
        "Credit",
        "Running Balance",
        "Transaction Reference",
        "Customer Reference",
        "Cheque Number",
        "Remarks 1",
        "Remarks 2",
    ]
    ws.append(header)
    for i in range(3030):
        debit = "100.00" if i % 2 == 0 else None
        credit = None if i % 2 == 0 else "50.00"
        ws.append(
            [
                datetime.datetime(2025, 12, 15, 0, 0),
                "",
                "",
                f"Txn {i}",
                "",
                "",
                debit,
                credit,
                f"{1000 + i}.00",
                f"TR{i}",
                f"CR{i}",
                "",
                "",
                "",
            ]
        )
    ws.append([None, "", "", "Total Credits", "", "", None, "999999.00", "", "", "", "", "", ""])
    ws.append([None, "", "", "Total Debits", "", "", "888888.00", None, "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_december_2025_equity_layout_preamble_header_and_parse():
    raw = _december_2025_layout_workbook_bytes()
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active

    assert ws.cell(row=6, column=5).value == "Total Count: 3030"

    idx, header_row = scan_equity_excel_header(ws)
    assert idx == 6
    assert header_row.count(None) == 2
    assert _is_equity_excel(header_row) is True

    mapping = _normalise_equity_excel_columns(header_row)
    assert mapping["date"] == 0
    assert mapping["description"] == 3
    assert mapping["debit"] == 6
    assert mapping["credit"] == 7
    assert mapping["balance"] == 8
    assert mapping["reference"] == 9

    wb.close()

    rows, _, _ = parse_xlsx(raw, "doc-dec-2025", "KES")
    assert len(rows) == 3030


def test_december_2025_idempotent_parse():
    """Same workbook bytes parsed twice must yield identical rows — determinism guard."""
    raw = _december_2025_layout_workbook_bytes()
    rows_a, hash_a, _ = parse_xlsx(raw, "doc-dec-2025", "KES")
    rows_b, hash_b, _ = parse_xlsx(raw, "doc-dec-2025", "KES")
    assert hash_a == hash_b
    assert rows_a == rows_b
