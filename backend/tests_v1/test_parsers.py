import io
import unittest
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook

from backend.v1.parsing.xlsx_parser import (
    parse_xlsx,
    _is_equity_excel,
    _normalise_equity_excel_columns,
)
from backend.v1.parsing.csv_parser import parse_csv
from backend.v1.parsing.errors import InvalidSchemaError

EQUITY_JAN_XLSX = Path(
    "/Users/mbakswatu/Desktop/parity/sayuni/2025/Excel/"
    "Sassy Cosmetics - Equity Bank - 1180279761781 - Jan 2025.xlsx"
)
EQUITY_FEB_XLSX = Path(
    "/Users/mbakswatu/Desktop/parity/sayuni/2025/Excel/"
    "Sassy Cosmetics - Equity Bank - 1180279761781 - Feb 2025.xlsx"
)


def _make_wb(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["date", "amount", "description"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestDeterministicParsers(unittest.TestCase):
    def test_xlsx_same_file_same_hash(self):
        rows = [
            ["2024-01-01", "100.00", "A"],
            ["2024-01-02", "-50", "B"],
        ]
        content = _make_wb(rows)
        parsed1, hash1, _ = parse_xlsx(content, "doc-1", "USD")
        parsed2, hash2, _ = parse_xlsx(content, "doc-1", "USD")
        self.assertEqual(parsed1, parsed2)
        self.assertEqual(hash1, hash2)

    def test_xlsx_reorder_rows_same_hash(self):
        rows_a = [
            ["2024-01-01", "100", "A"],
            ["2024-01-02", "200", "B"],
        ]
        rows_b = deepcopy(rows_a)[::-1]
        content_a = _make_wb(rows_a)
        content_b = _make_wb(rows_b)
        parsed_a, hash_a, _ = parse_xlsx(content_a, "doc-1", "USD")
        parsed_b, hash_b, _ = parse_xlsx(content_b, "doc-1", "USD")
        self.assertEqual(parsed_a, parsed_b)
        self.assertEqual(hash_a, hash_b)

    def test_xlsx_amount_equivalence(self):
        rows = [
            ["2024-01-01", "100", "A"],
            ["2024-01-02", "100.00", "B"],
        ]
        content = _make_wb(rows)
        parsed, _, _ = parse_xlsx(content, "doc-1", "USD")
        cents = [r["signed_amount_cents"] for r in parsed]
        self.assertIn(10000, cents)

    def test_xlsx_invalid_schema_missing_header(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["date", "description"])  # missing amount
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(InvalidSchemaError):
            parse_xlsx(buf.getvalue(), "doc-1", "USD")

    def test_equity_transacti_on_date_header_and_newline_in_date_cell(self):
        """Dec 2024 variant header + embedded newline in date string (Equity Excel)."""
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "Narrative",
                "Transacti on Date",
                "Debit",
                "Credit",
                "Running Balance",
            ]
        )
        ws.append(["Test", "02-12-\n2024", 100, None, None])
        buf = io.BytesIO()
        wb.save(buf)
        rows, _, _ = parse_xlsx(buf.getvalue(), "doc-dec", "KES")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["txn_date"], "2024-12-02")

    def test_csv_reorder_rows_same_hash(self):
        content_a = "date,amount,description\n2024-01-01,100,A\n2024-01-02,200,B\n"
        content_b = "date,amount,description\n2024-01-02,200,B\n2024-01-01,100,A\n"
        parsed_a, hash_a, _ = parse_csv(content_a.encode(), "doc-1", "USD")
        parsed_b, hash_b, _ = parse_csv(content_b.encode(), "doc-1", "USD")
        self.assertEqual(parsed_a, parsed_b)
        self.assertEqual(hash_a, hash_b)

    @unittest.skipUnless(EQUITY_JAN_XLSX.exists(), "January Equity Excel fixture missing")
    def test_equity_excel_january_mapping(self):
        from openpyxl import load_workbook

        wb = load_workbook(EQUITY_JAN_XLSX, data_only=True)
        ws = wb.worksheets[0]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertTrue(_is_equity_excel(header))
        mapping = _normalise_equity_excel_columns(header)
        for col in ("date", "description", "debit", "credit", "balance"):
            self.assertIn(col, mapping)

        rows, _, _ = parse_xlsx(EQUITY_JAN_XLSX.read_bytes(), "doc-jan", "KES")
        self.assertEqual(len(rows), 2677)
        signed = [r["signed_amount_cents"] for r in rows]
        self.assertTrue(any(v > 0 for v in signed))
        self.assertTrue(any(v < 0 for v in signed))

    @unittest.skipUnless(EQUITY_FEB_XLSX.exists(), "February Equity Excel fixture missing")
    def test_equity_excel_february_mapping(self):
        from openpyxl import load_workbook

        wb = load_workbook(EQUITY_FEB_XLSX, data_only=True)
        ws = wb.worksheets[0]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertTrue(_is_equity_excel(header))
        mapping = _normalise_equity_excel_columns(header)
        for col in ("date", "description", "debit", "credit", "balance"):
            self.assertIn(col, mapping)

        rows, _, _ = parse_xlsx(EQUITY_FEB_XLSX.read_bytes(), "doc-feb", "KES")
        self.assertGreater(len(rows), 0)


if __name__ == "__main__":
    unittest.main()
